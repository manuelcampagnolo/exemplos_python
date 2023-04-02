# https://stackoverflow.com/questions/42344041/how-to-copy-worksheet-from-one-workbook-to-another-one-using-openpyxl

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import os

# folder not necessary when working in a folder
#folder=r"C:\Users\mlc\OneDrive - Universidade de Lisboa\Documents\profissional-isa-cv\cg-isa"
#pathIn=os.path.join(folder,"DSD_2324_.xlsx")
#pathOut=os.path.join(folder,"DSD_2324_new.xlsx")
fnIn="DSD_2324_.xlsx"
fnOut="DSD_2324_new.xlsx"

############################################################ funcoes
def copy_sheet(source_sheet, target_sheet, idx):
    copy_cells(source_sheet, target_sheet, idx)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    if isinstance(source_sheet, openpyxl.worksheet._read_only.ReadOnlyWorksheet):
        return
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet,idx):
    delta=0
    for r, row in enumerate(source_sheet.iter_rows()):
        print('main r',r)
        # linhas em que há docente indicado
        if source_sheet.cell(r+1,idx+1).value: 
            print(source_sheet.cell(r+1,idx+1).value)
            copy_row(r+delta,row, source_sheet, target_sheet)
            if source_sheet.cell(r+1,idx+1).value!=column_key: # != 'docentes ' para evitar a 1a linha
                for k in range(N):
                    delta+=1
                    partial_copy_row(r+delta,row, source_sheet, target_sheet, columns_to_copy,coluna_validacao) # row comes from source; r+delta is the index in target
        else:
            copy_row(r+delta,row, source_sheet, target_sheet, fill=False)
        #if r>4: break
    print('delta: ',delta)


def copy_row(r,row, source_sheet, target_sheet, fill=True):
    for c, cell in enumerate(row):
        source_cell = cell
        if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
            continue
        target_cell = target_sheet.cell(column=c+1, row=r+1) # indices am cell começam em 1 ...!!!
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            if fill: target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)
        if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

def partial_copy_row(r,row, source_sheet, target_sheet, columns_to_copy,coluna_validacao):
    #print('r',r)
    for c, cell in enumerate(row):
        source_cell = cell
        if headers[c] in columns_to_copy:  
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices em cell começam em 1 ...!!!
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
        if headers[c]==coluna_validacao:
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices em cell começam em 1 ...!!!
            target_cell._value = VALIDATION_VALUE
            target_cell.data_type='s'

# number max docentes a inserir por UC e informação a passar
N=10 # número máximo docentes
columns_to_copy='Nome' # designação UC
column_key='docentes ' # cuidado: tem um espaço a mais
colunas_derivadas=['total Horas x n turmas da UC', 'Somatório', 'Horas em falta']
coluna_validacao=column_key

# worksheet
w=0
# source
wb_source = openpyxl.load_workbook(fnIn)#, data_only=True) # with data_only, it will only read values
wsnames=wb_source.sheetnames
source_sheet = wb_source[wsnames[w]]
print(source_sheet.max_row)
print(source_sheet.max_column)

# qual é a coluna "docentes "?
headers = [c.value for c in next(source_sheet.iter_rows(min_row=1, max_row=1))]
idx=headers.index(column_key)
VALIDATION_VALUE='Escolher docente'

# target work book and sheet
wb_target = openpyxl.Workbook()
target_sheet = wb_target.create_sheet(wsnames[w])

# create drop-down validation list
# lista de docentes para validação (drop-down menu)
val_sheet=wb_source['docentes']
mylist = [c.value for c in val_sheet['A']]
nomes_docentes=list(map(lambda x:x.strip(),mylist)) # eliminar \n
# escrever docentes em coluna AZ
for i in range(len(nomes_docentes)):
    target_sheet['AZ{}'.format(i+2)].value=nomes_docentes[i]
    #if i <5 : print(target_sheet['AZ{}'.format(i+2)].value)
#str1 = ','.join(nomes_docentes)                                    
#str1 = '"'+str1+'"'
#str1= '"Y,N"' # funciona!
#str1='"Vítor Manuel Delgado Alves,Docente a atribuir"' # funciona!
#str1='"Susete Maria Gonçalves Marques,Teresa de Jesus da Silva Matos Nolasco Crespo,Teresa Maria Gonçalves Quilhó Marques dos Santos,Teresa Paula Gonçalves Cruz (EU),Vítor Manuel Delgado Alves,Docente a atribuir"'
#data_val = DataValidation(type='list',formula1=str1) #, allow_blank=False)
data_val = DataValidation(type='list',formula1='=AZ:AZ') #, allow_blank=False)

# criar validação em target_sheet
target_sheet.add_data_validation(data_val)

# Criar cópia
copy_sheet(source_sheet, target_sheet,idx)

# depois de fazer a cópia:

# indicar as células em que ficam os drop-down menus
if True:
    # colocar drop-down menu para colunas em coluna_validacao (que têm valor VALIDATION_VALUE='VAL')
    idxval=headers.index(coluna_validacao)
    for k in range(target_sheet.max_row):
        c=target_sheet.cell(k+1,idxval+1)
        if c.value==VALIDATION_VALUE:
            data_val.add(c)
        #if k>5: break
#data_val.add(target_sheet["E5"]) #funciona!

# eliminar a worksheet original
if 'Sheet' in wb_target.sheetnames:  # remove default sheet
    wb_target.remove(wb_target['Sheet'])

# gravar para ficheiro
wb_target.save(fnOut)

wb_target.close


# what is this?
# isinstance(c, openpyxl.cell.read_only.EmptyCell)
