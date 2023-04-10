import pandas as pd
import os
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from datetime import datetime
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple


########################################################################### functions (não usado)

def df_to_excel(df, ws, header=True, index=True, startrow=0, startcol=0):
    """Write DataFrame df to openpyxl worksheet ws"""
    rows = dataframe_to_rows(df, header=header, index=index)
    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
             ws.cell(row=r_idx, column=c_idx).value = value

def create_workbook_from_dataframe(df):
    """
    1. Create workbook from specified pandas.DataFrame
    2. Adjust columns width to fit the text inside
    3. Make the index column and the header row bold
    4. Fill background color for the header row

    Other beautification MUST be done by usage side.
    """
    workbook = openpyxl.Workbook()
    ws = workbook.active

    rows = dataframe_to_rows(df.reset_index(), index=False)
    col_widths = [0] * (len(df.columns) + 1)
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):

            if type(val) is str:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            elif hasattr(val, "sort"):
                cell = ws.cell(row=i, column=j, value=", ".join(list(map(lambda v: str(v), list(val)))))
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            else:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val)) + 1])

            # Make the index column and the header row bold
            if i == 1 or j == 1:
                cell.font = Font(bold=True)

    # Adjust column width
    for i, w in enumerate(col_widths):
        letter = get_column_letter(i + 1)
        ws.column_dimensions[letter].width = w

    return workbook



############################################################################ inputs
xls_alunos_geom="students_1696_Geomática.2023-04-07_06-45-03_472.xlsx"
xls_alunos_sigdr="students_1761_Sistemas_de_Informação_Geográfica_e_Detecção_Remota.2023-04-07_06-45-28_144.xlsx"
xls_presencas="presencas_2022_2023_v2.xlsx"

# resultados 2022
df_2022=pd.read_csv('Pauta_apos_3a_chamada_6_julho_2022.csv',delimiter=';', encoding = 'unicode_escape')
df_2022.columns
vars_2022=['Número', 'Trab_2019-2020', 'Trab_2020-2021','Trab_2021_2022', 'Frequência']

# presenças/frequência 2022-2023
x1 = openpyxl.load_workbook(xls_presencas)
data = x1.active.values
columns = next(data)[0:]
df_pres=pd.DataFrame(data, columns=columns)
#df_pres=pd.read_csv('presencas_2022_2023_v2_10mar.csv',delimiter=';', encoding = 'unicode_escape')
df_pres.columns
vars_pres=['Número', 'Turma', 'GRUPO','Unidade 1', 'unidade 2', 'unidade 3']
df_pres['Unidade 1'].fillna('0', inplace=True)
print('unique values: ',df_pres['Unidade 1'].unique())
vals_unidade_sim=['sim', 'Sim','1']
[x for x in df_pres['Unidade 1'] if x in vals_unidade_sim]
vals_unidade_nao=['\xa0','0']
[x for x in df_pres['Unidade 1'] if x in vals_unidade_nao]

# ler lista alunos fenix 2022-2023
x1 = openpyxl.load_workbook(xls_alunos_geom) #, encoding='iso-8859-1')
x2 = openpyxl.load_workbook(xls_alunos_sigdr) #, encoding='iso-8859-1')
# converter para dataframe
data = x1.active.values
# Get the first line in file as a header line
columns = next(data)[0:]
# Create a DataFrame based on the second and subsequent lines of data
df1=pd.DataFrame(data, columns=columns)
data = x2.active.values
columns = next(data)[0:]
df2=pd.DataFrame(data, columns=columns)
df_insc=pd.concat([df1, df2], axis=0) 
df_insc['Número'] = df_insc['Número'].astype('int')
vars_insc=['Número', 'Nome','Estatutos']

############################################################################################
# criar tabela com todos os alunos inscritos com colunas:
# ['Número', 'Nome','Estatutos',"ano_frequencia", "nota_trabalho", unidade1, unidade2, unidade3, turma, grupo]
df_insc.columns
df_2022.columns
df=df_insc[vars_insc].merge(df_2022[vars_2022], how='left',on='Número')
df=df.merge(df_pres[vars_pres], how='left',on='Número')
df['Turma']=df['Turma'].astype('string')
# exportar para Excel
df.fillna("-",inplace=True)
df=df.sort_values(by='Turma')#.to_excel('aux.xlsx')
for var in ['Unidade 1', 'unidade 2', 'unidade 3']:
    df[var].replace(1, 'sim', inplace=True)
    df[var].replace('1', 'sim', inplace=True)
    df[var].replace('SIM', 'sim', inplace=True) 
#https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1772
# convert to string because df. applymap(str) do not like cetegories

df=df. applymap(str) 
df = df.reset_index(drop=True) # remove index column

wb=create_workbook_from_dataframe(df)

# change sheet name?


'''
# formatar output
wb = openpyxl.Workbook()
namews=datetime.today().strftime('%Y-%m-%d')
wb.create_sheet(namews)
ws=wb[namews]

for row in dataframe_to_rows(df, header = True, index=False):
    ws.append(row)

# headers
[c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

ws.column_dimensions['Nome'].width = 8
ws.column_dimensions['Estatutos'].width = 5
ws.column_dimensions['Frequência'].width = 12
for colname in ['Trab_2019-2020', 'Trab_2020-2021','Trab_2021_2022']:
    ws.column_dimensions[colname].width = len('Trab_2019-2020')


if 'Sheet' in wb.sheetnames:  # remove default sheet
    wb.remove(wb['Sheet'])
'''

fnout='alunos_estado_'+datetime.today().strftime('%Y-%m-%d')+'_nao_editar.xlsx'
wb.save(fnout)
